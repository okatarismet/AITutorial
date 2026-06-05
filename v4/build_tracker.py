# -*- coding: utf-8 -*-
"""Build a styled workout tracker .xlsx for Google Sheets upload."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---- palette ----
DARK   = "1F2A44"   # header navy
ACCENT = "2E5BFF"   # blue accent
LIGHT  = "EAEFFB"   # light blue row
GREEN  = "1E7F4F"
GREENL = "DDF2E6"
GREY   = "F4F6FA"
WHITE  = "FFFFFF"
ORANGE = "E8833A"

def fill(c):  return PatternFill("solid", fgColor=c)
def font(sz=11, b=False, c="222222"): return Font(name="Calibri", size=sz, bold=b, color=c)
thin = Side(style="thin", color="C9D2E3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

wb = Workbook()

# =========================================================================
# DASHBOARD
# =========================================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
widths = [3, 26, 16, 16, 16, 16, 3]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("B2:F2")
t = ws["B2"]; t.value = "NATURAL HYPERTROPHY  —  PROGRESS TRACKER"
t.font = font(18, True, WHITE); t.fill = fill(DARK); t.alignment = center
ws.row_dimensions[2].height = 38

ws.merge_cells("B3:F3")
s = ws["B3"]; s.value = "Upper / Lower · 4 days/week · each muscle 2× · every set 0–3 reps from failure"
s.font = font(11, False, WHITE); s.fill = fill(ACCENT); s.alignment = center
ws.row_dimensions[3].height = 22

# --- NEXT WORKOUT panel ---
ws.merge_cells("B5:F5")
h = ws["B5"]; h.value = "▶  WHAT'S NEXT"
h.font = font(13, True, WHITE); h.fill = fill(GREEN); h.alignment = left
ws.row_dimensions[5].height = 26

# helper cell to set the last completed day, drives the "next" formula
ws["B7"] = "Last workout I completed:"
ws["B7"].font = font(11, True)
dv = DataValidation(type="list", formula1='"Upper A,Lower A,Upper B,Lower B"', allow_blank=True)
ws.add_data_validation(dv)
ws["C7"] = "Lower B"
ws["C7"].font = font(11, True, c=ACCENT); ws["C7"].alignment = center
ws["C7"].fill = fill(LIGHT); ws["C7"].border = border
dv.add(ws["C7"])

ws.merge_cells("B9:F9")
nx = ws["B9"]
nx.value = ('=IF(C7="Upper A","➡  NEXT: LOWER A   (Squat · RDL · Leg Press · Calf · Abs)",'
            'IF(C7="Lower A","➡  NEXT: UPPER B   (Flat Press/Dip · Row · Pulldown · OHP · Laterals · Arms)",'
            'IF(C7="Upper B","➡  NEXT: LOWER B   (Hack Squat · RDL · Leg Curl · Calf · Abs)",'
            'IF(C7="Lower B","➡  NEXT: UPPER A   (Incline Press · Pull-up · OHP · Row · Laterals · Arms)",'
            '"Pick your last workout above ⤴"))))')
nx.font = font(14, True, DARK); nx.fill = fill(GREENL); nx.alignment = center
nx.border = border
ws.row_dimensions[9].height = 40
for r in (7,):
    for col in "BCDEF":
        ws[f"{col}{r}"].border = border

# --- Weekly schedule mini-table ---
ws.merge_cells("B11:F11")
sh = ws["B11"]; sh.value = "WEEKLY SCHEDULE"
sh.font = font(12, True, WHITE); sh.fill = fill(DARK); sh.alignment = left
ws.row_dimensions[11].height = 22

sched = [("Mon","Upper A"),("Tue","Lower A"),("Wed","Rest"),
         ("Thu","Upper B"),("Fri","Lower B"),("Sat","Rest"),("Sun","Rest")]
r = 12
for day, wk in sched:
    ws[f"B{r}"] = day; ws[f"B{r}"].font = font(11, True)
    ws.merge_cells(f"C{r}:F{r}")
    ws[f"C{r}"] = wk
    rest = wk == "Rest"
    ws[f"C{r}"].font = font(11, False, "8A93A5" if rest else DARK)
    ws[f"C{r}"].fill = fill(GREY if rest else LIGHT)
    ws[f"C{r}"].alignment = left
    for col in "BCDEF":
        ws[f"{col}{r}"].border = border
    r += 1

# --- How to use ---
ws.merge_cells(f"B{r+1}:F{r+1}")
g = ws[f"B{r+1}"]; g.value = "HOW TO USE"
g.font = font(12, True, WHITE); g.fill = fill(ORANGE); g.alignment = left
ws.row_dimensions[r+1].height = 22
tips = [
    "1.  Open the tab for today's workout (Upper A / Lower A / Upper B / Lower B).",
    "2.  Log Weight × Reps for every set. The sheet auto-flags if you beat last week.",
    "3.  Double progression: hit the TOP of the rep range on all sets → add 2.5–5 lb next time.",
    "4.  Keep every set 0–3 reps from failure. Beat the logbook each week — that's the whole game.",
    "5.  Update 'Last workout I completed' above so WHAT'S NEXT stays correct.",
]
rr = r+2
for tip in tips:
    ws.merge_cells(f"B{rr}:F{rr}")
    ws[f"B{rr}"] = tip; ws[f"B{rr}"].font = font(10.5); ws[f"B{rr}"].alignment = left
    ws.row_dimensions[rr].height = 18
    rr += 1

# =========================================================================
# WORKOUT DAY TABS
# =========================================================================
days = {
    "Upper A": [
        ("Incline Dumbbell Press (30°)", 4, "6–10"),
        ("Weighted Pull-up / Lat Pulldown", 4, "6–10"),
        ("Seated Overhead Press (BB/DB)", 3, "6–10"),
        ("Chest-Supported Row", 3, "8–12"),
        ("Lateral Raise (DB/cable)", 4, "12–20"),
        ("Incline DB Curl", 3, "8–12"),
        ("Overhead Triceps Extension", 3, "10–15"),
    ],
    "Lower A": [
        ("Squat (high-bar) / Hack Squat", 4, "6–10"),
        ("Romanian Deadlift (RDL)", 3, "8–12"),
        ("Leg Press", 3, "10–15"),
        ("Seated/Standing Calf Raise", 4, "10–15"),
        ("Weighted Cable Crunch", 3, "10–15"),
        ("Hanging Leg Raise", 3, "10–15"),
    ],
    "Upper B": [
        ("Flat DB Press / Weighted Dip", 4, "8–12"),
        ("Chest-Supported Row", 4, "8–12"),
        ("Lat Pulldown", 3, "8–12"),
        ("Seated Overhead Press (BB/DB)", 3, "8–12"),
        ("Lateral Raise (DB/cable)", 4, "12–20"),
        ("Incline DB Curl", 3, "8–12"),
        ("Overhead Triceps Extension", 3, "10–15"),
    ],
    "Lower B": [
        ("Hack Squat / Squat", 4, "8–12"),
        ("Romanian Deadlift (RDL)", 3, "8–12"),
        ("Seated Leg Curl", 3, "10–15"),
        ("Seated/Standing Calf Raise", 4, "12–20"),
        ("Weighted Decline Crunch / Ab Wheel", 3, "8–15"),
        ("Hanging Leg Raise", 3, "10–15"),
    ],
}

SESSIONS = 8  # repeated session blocks stacked vertically (mobile scroll-down)

def build_day(name, rows):
    """Mobile-first VERTICAL layout: narrow columns, scroll DOWN through
    repeated session blocks. No sideways scrolling on a phone."""
    sh = wb.create_sheet(name)
    sh.sheet_view.showGridLines = False
    # Narrow columns so the whole table fits a phone's width.
    # A pad | B Exercise | C Target | D Weight | E Reps | F Done
    sh.column_dimensions["A"].width = 1.5
    sh.column_dimensions["B"].width = 24
    sh.column_dimensions["C"].width = 7
    sh.column_dimensions["D"].width = 8
    sh.column_dimensions["E"].width = 6
    sh.column_dimensions["F"].width = 5
    LAST = 6  # column F

    # --- Sticky title (frozen) ---
    sh.merge_cells(start_row=1, start_column=2, end_row=1, end_column=LAST)
    t = sh.cell(1, 2, f"{name.upper()}")
    t.font = font(16, True, WHITE); t.fill = fill(DARK); t.alignment = center
    sh.row_dimensions[1].height = 30

    sh.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LAST)
    s = sh.cell(2, 2, "Scroll down · log each session · 0–3 reps from failure")
    s.font = font(9.5, False, WHITE); s.fill = fill(ACCENT); s.alignment = center
    sh.row_dimensions[2].height = 16

    r = 3
    for sess in range(SESSIONS):
        # Session date bar
        sh.merge_cells(start_row=r, start_column=2, end_row=r, end_column=LAST)
        d = sh.cell(r, 2, f"  SESSION {sess+1}   ·   Date: ____ / ____")
        d.font = font(11, True, WHITE)
        d.fill = fill(GREEN if sess == 0 else ORANGE)
        d.alignment = left; d.border = border
        sh.row_dimensions[r].height = 22
        r += 1

        # Column headers for this block
        hdrs = ["Exercise", "Target", "Weight", "Reps", "✓"]
        for i, h in enumerate(hdrs):
            c = sh.cell(r, 2+i, h)
            c.font = font(9.5, True, WHITE); c.fill = fill(DARK)
            c.alignment = center; c.border = border
        sh.row_dimensions[r].height = 16
        r += 1

        # Exercise rows
        for i, (ex, sets, reps) in enumerate(rows):
            zebra = LIGHT if i % 2 == 0 else WHITE
            target = f"{sets}×{reps}"
            ce = sh.cell(r, 2, ex); ce.font = font(10.5); ce.alignment = left; ce.fill = fill(zebra)
            ct = sh.cell(r, 3, target); ct.font = font(10.5, True, ACCENT); ct.alignment = center; ct.fill = fill(zebra)
            cw = sh.cell(r, 4, None); cw.fill = fill(zebra); cw.alignment = center; cw.font = font(11)
            cr = sh.cell(r, 5, None); cr.fill = fill(zebra); cr.alignment = center; cr.font = font(11)
            cd = sh.cell(r, 6, None); cd.fill = fill(zebra); cd.alignment = center; cd.font = font(11)
            for col in range(2, 7):
                sh.cell(r, col).border = border
            sh.row_dimensions[r].height = 21
            r += 1

        # Small spacer between sessions
        sh.row_dimensions[r].height = 6
        r += 1

    # Freeze the title + subtitle so the day name stays on screen while scrolling
    sh.freeze_panes = "A3"
    return sh

for name, rows in days.items():
    build_day(name, rows)

# =========================================================================
# BODYWEIGHT / PROGRESS tab
# =========================================================================
bw = wb.create_sheet("Bodyweight & Notes")
bw.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [2, 14, 14, 14, 40]):
    bw.column_dimensions[col].width = w
bw.merge_cells("B2:E2")
t = bw["B2"]; t.value = "BODYWEIGHT  &  WEEKLY NOTES"
t.font = font(16, True, WHITE); t.fill = fill(DARK); t.alignment = center
bw.row_dimensions[2].height = 30
heads = ["Date", "Bodyweight", "Sleep (avg h)", "Notes (energy, pain, deload, diet)"]
for i, h in enumerate(heads):
    c = bw.cell(4, 2+i, h); c.font = font(11, True, WHITE); c.fill = fill(ACCENT)
    c.alignment = center; c.border = border
bw.row_dimensions[4].height = 20
for r in range(5, 33):
    zebra = LIGHT if r % 2 == 1 else WHITE
    for col in range(2, 6):
        cc = bw.cell(r, col); cc.fill = fill(zebra); cc.border = border
        cc.alignment = left if col == 5 else center; cc.font = font(11)
    bw.row_dimensions[r].height = 19

# save
out = r"C:\Users\STAJYER\Desktop\ismet_claude_projects\AITutorial\v4\workout-tracker.xlsx"
wb.save(out)
print("saved:", out)
